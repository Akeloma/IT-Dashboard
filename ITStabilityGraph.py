import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook

def main():
    st.title("📊 Monthly IT Stability Data Appender")
    st.markdown("Upload **8 source files** (with `Page 1` sheets) and **1 target file** (`All Data.xlsx`).")

    # --- Multi-file upload for sources ---
    source_files = st.file_uploader(
        "📂 Upload ALL 8 Source Files (.xlsx)",
        type=["xlsx"],
        accept_multiple_files=True,
        key="sources"
    )

    # --- Single target file uploader ---
    target_file = st.file_uploader("📂 Upload Target File (.xlsx)", type=["xlsx"], key="target")

    if source_files and target_file:
        sheet_source = "Page 1"
        sheet_target = "IT Stability"
        cols_to_extract = [0, 1, 2, 5, 14, 7, 8, 9, 11]  # A,B,C,F,O,H,I,J,K

        try:
            # --- STEP 1: Extract and combine all source files ---
            extracted_list = []
            for uploaded in source_files:
                df = pd.read_excel(uploaded, sheet_name=sheet_source)
                selected = df.iloc[:, cols_to_extract]
                extracted_list.append(selected)

            combined_sources = pd.concat(extracted_list, ignore_index=True)
            st.success(f"✅ Extracted data from {len(source_files)} files.")
            st.dataframe(combined_sources.head())

            # --- STEP 2: Read target workbook ---
            target_df = pd.read_excel(target_file, sheet_name=sheet_target)

            # Merge + overwrite (fast mode)
            combined_df = pd.concat([target_df, combined_sources], ignore_index=True)

            # --- STEP 3: Write to BytesIO ---
            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl", mode="w") as writer:
                combined_df.to_excel(writer, index=False, sheet_name=sheet_target)
            output.seek(0)

            # --- STEP 4: Apply formatting ---
            wb = load_workbook(output)
            ws = wb[sheet_target]

            # Format date columns (H & I)
            for col in ["H", "I"]:
                for cell in ws[col][1:]:
                    if cell.value not in (None, "", "NaT"):
                        cell.number_format = "mmm-yy"

            # Apply column padding
            col_widths = {
                "A": 15, "B": 20, "C": 18, "D": 25, "E": 25,
                "F": 22, "G": 30, "H": 14, "I": 14
            }
            for col, width in col_widths.items():
                ws.column_dimensions[col].width = width

            new_output = BytesIO()
            wb.save(new_output)
            wb.close()

            st.success(f"🗓️ Appended {len(combined_sources)} rows from {len(source_files)} files successfully!")

            # --- STEP 5: Download ---
            st.download_button(
                label="📥 Download Updated File",
                data=new_output.getvalue(),
                file_name="All_Data_Updated.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"❌ Error: {e}")

    else:
        st.info("👆 Upload all 8 source files + 1 target file to begin.")


if __name__ == "__main__":
    main()


