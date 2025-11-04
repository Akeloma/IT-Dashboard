import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from copy import copy
from datetime import datetime
from io import BytesIO
from openpyxl.utils.datetime import from_excel
from openpyxl.styles import Alignment

def main():
        st.header("📈 Toxic Streamlit - Upload output file in Sharepoint")
        st.write("""
        Upload your 'manual calculated' Excel and your new data file.
        The app will copy everything from **row 2 onwards** from the new data and paste it into
        the **'Overall database'** sheet (preserving formatting and adding borders).
        """)

        manual_file = st.file_uploader("📘 Upload 'manual calculated' Excel file", type=["xlsx"])
        new_file = st.file_uploader("📗 Upload new data file", type=["xlsx"])

        if manual_file and new_file:
            try:
                manual_file.seek(0)
                new_file.seek(0)
                wb_manual = load_workbook(manual_file)
                wb_new = load_workbook(new_file)

                ws_manual = wb_manual["Overall database"]
                ws_new = wb_new.active

                # --- FIND TRUE LAST FILLED ROW ---
                def get_last_filled_row(ws):
                    for row in range(ws.max_row, 0, -1):
                        if any(cell.value not in (None, "") for cell in ws[row]):
                            return row
                    return 1

                # --- ADD ONE MONTH ---
                def add_one_month(dt: datetime) -> datetime:
                    y, m = dt.year, dt.month
                    if m == 12:
                        return datetime(y + 1, 1, 1)
                    else:
                        return datetime(y, m + 1, 1)
      
                # --- FIND 'File' & 'Date' COLUMNS (auto-detect header row) ---
                header_row_idx = None
                file_col = None
                date_col = None

                # Search first 10 rows to find where the header row actually is
                for r in range(1, 11):
                    row_values = [str(c.value).strip().lower() if c.value else "" for c in ws_manual[r]]
                    if "file" in row_values and "date" in row_values:
                        header_row_idx = r
                        break

                if header_row_idx:
                    header_map = {str(cell.value).strip().lower(): idx + 1 for idx, cell in enumerate(ws_manual[header_row_idx]) if cell.value}
                    file_col = header_map.get("file", None)
                    date_col = header_map.get("date", None)

                st.write(f"🧭 Header row detected at: {header_row_idx}, File col: {file_col}, Date col: {date_col}")

                # --- FIND LAST DATE VALUE IN FILE OR DATE COLUMN ---
                def get_last_date_value(ws, col_idx):
                    """Return the last valid date from a column, supporting datetime, strings, and Excel serials."""
                    if not col_idx:
                        return None

                    for r in range(ws.max_row, 1, -1):
                        val = ws.cell(row=r, column=col_idx).value
                        if val in (None, ""):
                            continue

                        # Case 1: Already a datetime object
                        if isinstance(val, datetime):
                            return val

                        # Case 2: Excel serial date (float or int)
                        if isinstance(val, (float, int)):
                            try:
                                return from_excel(val)
                            except Exception:
                                continue

                        # Case 3: String-formatted date
                        s = str(val).strip()
                        for fmt in ("%d-%b-%y", "%d-%b-%Y", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d"):
                            try:
                                return datetime.strptime(s, fmt)
                            except Exception:
                                continue

                    return None

                # --- DETECT LAST DATE AND ADD ONE MONTH (robust version) ---
                last_row_idx = get_last_filled_row(ws_manual)
                base_date = None

                # Search upwards for the last non-empty File or Date cell
                if file_col or date_col:
                    search_col = (file_col or date_col) or ws_manual.max_column
                    for r in range(last_row_idx, 1, -1):
                        val = ws_manual.cell(row=r, column=search_col).value
                        if val not in (None, ""):
                            # Handle Excel serial, datetime, or text-formatted date
                            try:
                                if isinstance(val, datetime):
                                    base_date = val
                                elif isinstance(val, (int, float)):
                                    from openpyxl.utils.datetime import from_excel
                                    base_date = from_excel(val)
                                else:
                                    s = str(val).strip()
                                    for fmt in ("%d-%b-%y", "%d-%b-%Y", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d"):
                                        try:
                                            base_date = datetime.strptime(s, fmt)
                                            break
                                        except Exception:
                                            continue
                                if base_date:
                                    break
                            except Exception:
                                continue

                # If still nothing found, default to current date
                if not base_date:
                    base_date = datetime.now()

                new_month_date = add_one_month(base_date)

                # Format to '1-Oct-25' style
                try:
                    new_month_str = new_month_date.strftime("%-d-%b-%y")
                except Exception:
                    new_month_str = new_month_date.strftime("%#d-%b-%y")

                st.write(f"📅 Detected last date: {base_date}, next month: {new_month_str}")

                # --- DEFINE BORDER STYLE ---
                thin_border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin")
                )

                # --- DEFINE COPY RANGE ---
                start_row = 2
                end_row = ws_new.max_row
                max_col = ws_new.max_column
                last_row = get_last_filled_row(ws_manual)

                # --- COPY ROWS (SKIP BLANK ONES) ---
                for r in range(start_row, end_row + 1):
                    if not any(ws_new.cell(row=r, column=c).value not in (None, "") for c in range(1, max_col + 1)):
                        continue

                    last_row += 1
                    for c in range(1, max_col + 1):
                        src_cell = ws_new.cell(row=r, column=c)
                        tgt_cell = ws_manual.cell(row=last_row, column=c)
                        tgt_cell.value = src_cell.value
                        if src_cell.has_style:
                            tgt_cell.font = copy(src_cell.font)
                            tgt_cell.fill = copy(src_cell.fill)
                            tgt_cell.alignment = copy(src_cell.alignment)
                            tgt_cell.alignment = tgt_cell.alignment.copy(wrap_text=False)
                            tgt_cell.number_format = src_cell.number_format
                        tgt_cell.border = thin_border

                    # --- Fill File & Date columns ---
                    if file_col:
                        fcell = ws_manual.cell(row=last_row, column=file_col)
                        fcell.value = new_month_date        # store actual datetime
                        fcell.number_format = "d-mmm-yy"    # show as 1-Sep-25
                        fcell.border = thin_border
                        fcell.alignment = Alignment(horizontal="left", vertical="center")  # 👈 added

                    if date_col:
                        dcell = ws_manual.cell(row=last_row, column=date_col)
                        dcell.value = new_month_date
                        dcell.number_format = "d-mmm-yy"
                        dcell.border = thin_border
                        dcell.alignment = Alignment(horizontal="left", vertical="center")  # 👈 added


                # --- SAVE OUTPUT ---
                output = BytesIO()
                wb_manual.save(output)
                output.seek(0)

                st.success("✅ Data pasted successfully into 'Overall database'!")
                st.download_button(
                    label="⬇️ Download Combined File",
                    data=output,
                    file_name=f"manual_calculated_combined_{datetime.now().strftime('%d%b%y')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            except Exception as e:
                st.error(f"❌ An error occurred: {e}")

        else:
            st.info("👆 Please upload both Excel files to begin.")
