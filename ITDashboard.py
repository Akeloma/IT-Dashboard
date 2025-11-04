import streamlit as st
import pandas as pd
import re
from io import BytesIO
from openpyxl import load_workbook
from datetime import datetime
from dateutil.relativedelta import relativedelta

# === Helper Functions ===
def clean_oe_name(name: str) -> str:
    if not isinstance(name, str):
        return ""
    # normalize spaces and dashes but KEEP Ⓖ
    s = name.replace("\u00A0", " ")  # remove non-breaking space
    s = re.sub(r"[–—]", "-", s)      # normalize dashes
    s = re.sub(r"\s+", " ", s).strip()
    return s

def normalize_columns(df):
    """Standardize column names (remove NBSPs, normalize dashes/spaces)."""
    df.columns = (
        df.columns
        .str.replace("\u00A0", " ", regex=True)
        .str.replace("–", "-", regex=True)
        .str.strip()
    )
    return df

def parse_sheet(raw_bytes, sheet_name, wanted_headers):
    """
    Extract KPI data from a sheet (merged-cell friendly).
    Returns list of dicts: [{'OE': 'Allianz Malaysia', 'KPI1': 'Good (99.9)', ...}]
    """
    wb = load_workbook(raw_bytes, data_only=True)
    ws = wb[sheet_name]

    header_map = {}
    for r in range(1, 30):
        for c in range(1, 25):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                vv = v.strip().replace("–", "-")
                for hdr in wanted_headers:
                    if hdr.lower() in vv.lower():
                        header_map[c] = hdr  # use canonical header name

    if not header_map:
        raise ValueError(f"❌ Could not find KPI headers in sheet: {sheet_name}")

    rating_re = re.compile(r"(Very Bad|Bad|Medium|Good|Very Good)\s*\(([-+]?\d+(\.\d+)?)\)")
    results = []
    for r in range(20, 100):
        oe_name = ws.cell(row=r, column=5).value  # col E
        if not isinstance(oe_name, str):
            continue
        oe_name_clean = clean_oe_name(oe_name)
        if not oe_name_clean:
            continue

        row_data = {"OE": oe_name_clean}
        found = False
        for c, hdr in header_map.items():
            val = ws.cell(row=r, column=c).value
            if isinstance(val, str) and rating_re.search(val):
                row_data[hdr] = val.strip()
                found = True
            else:
                row_data.setdefault(hdr, "")
        if found:
            results.append(row_data)
    return pd.DataFrame(results)

# === Streamlit UI ===
def main():     
        st.header("🏗️ ITIS Cockpit – Upload output file in Sharepoint")

        st.markdown("""
        Upload:
        1️⃣ **Raw Data.xlsx** (Archer file, with MY data filled)  
        2️⃣ **KPI2 Excel file** (download latest one from Sharepoint)  
        This gives the output file to be processed in Power BI.
        """)

        raw_file = st.file_uploader("📥 Upload Raw Data Excel", type=["xlsx"])
        kpi_file = st.file_uploader("📥 Upload KPI2 Excel", type=["xlsx"])

        # === MAIN LOGIC ===
        if raw_file and kpi_file:
            try:
                # Step 1 – Load KPI2
                kpi_df = pd.read_excel(kpi_file, sheet_name="KPI2")
                kpi_df["OE_clean"] = kpi_df["OE"].astype(str).apply(clean_oe_name)
                kpi_df["OE_clean_nog"] = kpi_df["OE_clean"].str.replace("Ⓖ", "", regex=False).str.strip()

                kpi_df = normalize_columns(kpi_df)

                # Step 2 – Get next month (keep as datetime)
                last_date_raw = kpi_df["Date"].dropna().iloc[-1]
                last_date = datetime.strptime(last_date_raw, "%b-%y") if isinstance(last_date_raw, str) else pd.to_datetime(last_date_raw)
                next_month = last_date + relativedelta(months=1)
                month_label = next_month  # store as datetime, not text

                # Step 3 – Define sheets and KPIs
                sheet_kpi_map = {
                    "IT Strategy & Project Port(H06)": [
                        "Group IT Strategy Alignment Score", "IT Steering Board Score"
                    ],
                    "Architecture & Demand(H06)": [
                        "Architecture Data Quality Score", "Cloudification",
                        "Information Domain - Initialization", "Information Domain - Documentation", "Legacy Index"
                    ],
                    "IT Service Agreements & Su(H06)": [
                        "Group Toxicity", "Local Toxicity", "Overall Toxicity",
                        "IT Asset Lifecycle Management Score"
                    ],
                    "IT Governance, Risk & Comp(H06)": [
                        "IT Compliance Score", "Unmanaged Risks: ITOM", "Unmanaged Risks: ITOM+ISMS+BA",
                        "Completed risk scoping BAs", "Completed risk scoping BAs: regular", "Completed risk scoping BAs: EUCs"
                    ]
                }

                # Step 4 – Parse and merge horizontally (by OE)
                merged_df = None
                for sheet, kpis in sheet_kpi_map.items():
                    st.write(f"🔍 Reading **{sheet}** ...")
                    temp_df = parse_sheet(raw_file, sheet, kpis)
                    if merged_df is None:
                        merged_df = temp_df
                    else:
                        merged_df = pd.merge(merged_df, temp_df, on="OE", how="outer")

                # Step 5 – Create final append_df
                append_rows = []
                for _, rec in merged_df.iterrows():
                    oe_clean = clean_oe_name(rec["OE"])
                    # Try match including Ⓖ first, then fallback to version without it
                    cim_id_val = kpi_df.loc[kpi_df["OE_clean"].str.lower() == oe_clean.lower(), "CIM ID / OE ID"]
                    if cim_id_val.empty:
                        cim_id_val = kpi_df.loc[
                            kpi_df["OE_clean_nog"].str.lower()
                            == oe_clean.lower().replace("Ⓖ", "").strip(),
                            "CIM ID / OE ID"
                        ]
                    cim_id_val = cim_id_val.iloc[0] if not cim_id_val.empty else ""

                    row_dict = {"Date": month_label, "CIM ID / OE ID": cim_id_val, "OE": oe_clean}
                    for col in merged_df.columns:
                        if col != "OE":
                            row_dict[col] = rec[col]
                    append_rows.append(row_dict)


                append_df = pd.DataFrame(append_rows)
                append_df = normalize_columns(append_df)

                # Step 6 – Append below last month
                updated = pd.concat([kpi_df.drop(columns=["OE_clean", "OE_clean_nog"]), append_df], ignore_index=True)
                updated["OE"] = updated["OE"].astype(str).apply(clean_oe_name)

                # Step 7 – Sort OEs in custom order
                custom_order = [
                    "Allianz China - Holding",
                    "Allianz Indonesia",
                    "Allianz Philippine - L&H",
                    "Allianz SingaporeⒼ",
                    "Allianz Sri Lanka",
                    "Allianz Taiwan - Life",
                    "Allianz Thailand",
                    "Allianz Australia - P&CⒼ",
                    "Allianz Malaysia"
                ]
                # ✅ Convert OE to category first (important!)
                updated["OE"] = pd.Categorical(updated["OE"], categories=custom_order, ordered=True)

                # ✅ Sort by Date first, then OE order
                updated = updated.sort_values(["Date", "OE"], ascending=[True, True], key=lambda col: (
                    col.map(lambda x: x if isinstance(x, pd.Timestamp) else pd.to_datetime(x, errors="coerce"))
                    if col.name == "Date"
                    else col.cat.codes
                )).reset_index(drop=True)


                # Step 8 – Output
                st.success(f"✅ Added {len(append_df)} new rows for {month_label.strftime('%b-%y')} (side-by-side, aligned).")
                st.dataframe(append_df)

                src_wb = load_workbook(kpi_file)
                if "Sheet1" in src_wb.sheetnames:
                    src_ws = src_wb["Sheet1"]
                    sheet1_data = pd.DataFrame(src_ws.values)
                else:
                    sheet1_data = None
                buf = BytesIO()

                with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                    updated.to_excel(writer, index=False, sheet_name="KPI2")
                    ws = writer.sheets["KPI2"]
                    for column_cells in ws.iter_cols(min_col=1, max_col=1):
                        for cell in column_cells:
                            cell.number_format = "mmm-yy"

                    if sheet1_data is not None:
                        sheet1_data.to_excel(writer, index=False, header=False, sheet_name="Sheet1")

                    # === Auto-adjust column widths for all sheets ===
                    for ws in writer.book.worksheets:
                        for column_cells in ws.columns:
                            max_length = 0
                            column = column_cells[0].column_letter
                            for cell in column_cells:
                                try:
                                    cell_value = str(cell.value)
                                    if len(cell_value) > max_length:
                                        max_length = len(cell_value)
                                except:
                                    pass
                            adjusted_width = max_length + 0.5  # add a little padding
                            ws.column_dimensions[column].width = adjusted_width


                st.download_button(
                    "💾 Download Updated KPI2.xlsx",
                    data=buf.getvalue(),
                    file_name="Updated_KPI2.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            except Exception as e:
                st.error(f"❌ Error: {e}")