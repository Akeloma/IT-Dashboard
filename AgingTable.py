import streamlit as st
import pandas as pd
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from io import BytesIO


# ------------------------------------------------
# CONFIG
# ------------------------------------------------
OE_ORDER = ["AZCH", "ID", "MY", "PH", "AIS", "SL", "TH", "TW", "AZAP"]
METRICS = [
    "Incidents Total Aging > 31 days",
    "Incidents Aging > 31-90 days",
    "Incidents Aging > 90 days",
]

RAW_SHEET_NAME = "Page 1"
RAW_CREATED_COL = "Created"
RAW_RESOLVED_COL = "Resolved"
RAW_OE_COL = "Affected OEs"

RAW_TO_OE_MAP = [
    ("Allianz China", "AZCH"),
    ("Allianz China - P&C", "AZCH"),
    ("Allianz Indonesia", "ID"),
    ("Allianz Malaysia", "MY"),
    ("Allianz Philippine", "PH"),
    ("Allianz Singapore", "AIS"),
    ("Allianz Sri Lanka", "SL"),
    ("Allianz Thailand", "TH"),
    ("Allianz Taiwan", "TW"),
    ("Allianz SE Singapore Branch OE", "AZAP"),
]

# ------------------------------------------------
# HELPER FUNCTIONS
# ------------------------------------------------
def normalize_oe(raw_val: str) -> str | None:
    if pd.isna(raw_val):
        return None
    s = str(raw_val).lower()
    if "," in s:
        s = s.split(",")[0].strip()
    for key_substr, code in RAW_TO_OE_MAP:
        if key_substr.lower() in s:
            return code
    return None


def compute_counts_from_raw(raw_file: str) -> pd.DataFrame:
    """
    Read the raw aging Excel and compute counts per (OE, Metric).
    Returns a DataFrame with columns: OE, Metric, Value
    Missing OEs/metrics are filled later with 0.
    """
    raw = pd.read_excel(raw_file, sheet_name=RAW_SHEET_NAME)
    raw.columns = [c.strip() for c in raw.columns]

    # Convert to datetime and include unresolved incidents (treated as today)
    raw[RAW_CREATED_COL] = pd.to_datetime(raw[RAW_CREATED_COL], errors="coerce")
    raw[RAW_RESOLVED_COL] = pd.to_datetime(raw[RAW_RESOLVED_COL], errors="coerce")
    raw["Days"] = (raw[RAW_RESOLVED_COL].fillna(datetime.today()) - raw[RAW_CREATED_COL]).dt.days
    raw = raw.dropna(subset=["Days"])

    # Normalize OEs
    raw["OE"] = raw[RAW_OE_COL].apply(normalize_oe)
    raw = raw.dropna(subset=["OE"])

    # --- Summary section ---
    print("🔎 OE normalization value counts:")
    print(raw["OE"].value_counts(dropna=False).head(20))

    print("\n📊 Days summary (after dropping NaNs):")
    print(raw["Days"].describe())

    # --- Over 31 days ---
    gt31 = raw[raw["Days"] > 30].groupby("OE").size().rename("Value").reset_index()
    print("\n🧮 Over 31 days per OE:")
    print(gt31.sort_values("Value", ascending=False))

    # --- 31–90 days ---
    in_31_90 = raw[(raw["Days"] >= 31) & (raw["Days"] <= 90)].groupby("OE").size().rename("Value").reset_index()
    print("\n📆 Between 31–90 days per OE:")
    print(in_31_90.sort_values("Value", ascending=False))

    # --- Over 90 days ---
    gt90 = raw[raw["Days"] > 90].groupby("OE").size().rename("Value").reset_index()
    print("\n⏰ Over 90 days per OE:")
    print(gt90.sort_values("Value", ascending=False))

    # --- Add metric labels ---
    gt31["Metric"] = "Incidents Total Aging > 31 days"
    in_31_90["Metric"] = "Incidents Aging > 31-90 days"
    gt90["Metric"] = "Incidents Aging > 90 days"

    # Combine all
    counts = pd.concat([gt31, in_31_90, gt90], ignore_index=True)
    return counts[["OE", "Metric", "Value"]]


def append_next_month_with_counts(powerbi_file, raw_file):
    existing = pd.read_excel(powerbi_file)
    last_date = pd.to_datetime(existing["Date"].iloc[-1])
    next_month = last_date + relativedelta(months=1)
    st.info(f"📅 Detected last date: {last_date.strftime('%b-%y')}, next month: {next_month.strftime('%b-%y')}")

    counts = compute_counts_from_raw(raw_file)
    key_to_val = {(row.OE, row.Metric): int(row.Value) for _, row in counts.iterrows()}

    # Build 27-row structure
    rows = []
    for metric in METRICS:
        for oe in OE_ORDER:
            val = key_to_val.get((oe, metric), 0)
            rows.append({"OE": oe, "Metric": metric, "Date": next_month, "Value": val})

    new_block = pd.DataFrame(rows)
    combined = pd.concat([existing, new_block], ignore_index=True)
    combined["Date"] = pd.to_datetime(combined["Date"])

    # --- Save to memory instead of disk ---
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        combined.to_excel(writer, index=False)
        ws = writer.sheets["Sheet1"]

        # Apply formatting
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
            for cell in row:
                cell.number_format = "mmm-yy"
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                cell_value = str(cell.value)
                if cell_value:
                    max_length = max(max_length, len(cell_value))
            ws.column_dimensions[col_letter].width = max_length + 1

    output.seek(0)
    st.success(f"✅ Appended {len(new_block)} rows for {next_month.strftime('%b-%y')}")
    return combined, output

# ------------------------------------------------
# STREAMLIT UI
# ------------------------------------------------
def main():
    # Only set page config if this file is run standalone
    if not st.runtime.exists():
        st.set_page_config(page_title="Aging Table Automation", layout="wide")
    st.title("📊 Aging Incident Table Updater")

    st.write("Upload your PowerBI ITSM file and the new aging incidents raw file.")
    col1, col2 = st.columns(2)

    with col1:
        powerbi_file = st.file_uploader("📘 PowerBI ITSM Excel", type=["xlsx"])
    with col2:
        raw_file = st.file_uploader("📗 Aging Incidents file", type=["xlsx"])

    if powerbi_file and raw_file:
        with st.spinner("Processing and updating data..."):
            df, output = append_next_month_with_counts(powerbi_file, raw_file)

        # --- Download Button ---
        st.download_button(
            label="📥 Download Updated PowerBI File",
            data=output,
            file_name="Updated_PowerBI_ITSM.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.info("👆 Please upload both files to continue.")
